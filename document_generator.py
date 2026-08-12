from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _sanitize_sheet_title(title: str) -> str:
    sanitized = title.replace("/", "_").replace("\\", "_").replace("[", "_").replace("]", "_").replace("*", "_").replace("?", "_").replace(":", "_")
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    return sanitized


def _set_header_style(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value or "")
            except Exception:
                value = ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)


def create_table_sheet(wb: Workbook, table_name: str, columns: List[Dict[str, Any]]) -> None:
    sheet_title = _sanitize_sheet_title(table_name)
    if sheet_title in wb.sheetnames:
        sheet_title = f"{sheet_title}_1"
    ws = wb.create_sheet(title=sheet_title)
    header = [
        "Column Name",
        "Data Type",
        "Nullable",
        "Primary Key",
        "Foreign Key",
        "Referenced Table",
        "Referenced Column",
        "Default Value",
        "Description",
    ]
    ws.append([f"Table: {table_name}"] + [""] * (len(header) - 1))
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in sorted(columns, key=lambda item: item.get("ordinal_position", 0)):
        fk = c.get("foreign_key") or {}
        ws.append(
            [
                c.get("column_name"),
                c.get("data_type"),
                "YES" if c.get("is_nullable") else "NO",
                "YES" if c.get("is_primary") else "",
                fk.get("constraint_name", "") if fk else "",
                fk.get("referenced_table", "") if fk else "",
                fk.get("referenced_column", "") if fk else "",
                c.get("column_default") or "",
                c.get("column_comment") or "",
            ]
        )
    ws.freeze_panes = "A3"
    _auto_size_columns(ws)


def create_table_index_sheet(wb: Workbook, schema_map: Dict[str, List[Dict[str, Any]]]) -> None:
    ws = wb.create_sheet(title="Table_Index")
    header = ["Table Name", "Columns", "Primary Key", "Foreign Keys", "Status"]
    ws.append(header)
    for cell in ws[1]:
        _set_header_style(cell)
    for table_name, columns in schema_map.items():
        pk_columns = [c["column_name"] for c in columns if c.get("is_primary")]
        fk_columns = [c["column_name"] for c in columns if c.get("foreign_key")]
        ws.append([
            table_name,
            len(columns),
            ", ".join(pk_columns) if pk_columns else "-",
            ", ".join(fk_columns) if fk_columns else "-",
            "Processed",
        ])
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def create_table_relationships_sheet(wb: Workbook, schema_map: Dict[str, List[Dict[str, Any]]]) -> None:
    ws = wb.create_sheet(title="Table_Relationships")
    header = ["Source Table", "Source Column", "Target Table", "Target Column", "Constraint Name"]
    ws.append(header)
    for cell in ws[1]:
        _set_header_style(cell)
    for table_name, columns in schema_map.items():
        for column in columns:
            fk = column.get("foreign_key")
            if fk:
                ws.append([
                    table_name,
                    column.get("column_name"),
                    fk.get("referenced_table", ""),
                    fk.get("referenced_column", ""),
                    fk.get("constraint_name", ""),
                ])
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def create_readme_sheet(wb: Workbook, connection_name: str, tables: List[str]) -> None:
    ws = wb.create_sheet(title="README")
    lines = [
        f"Database Schema Documentation",
        f"Connection: {connection_name}",
        "",
        "This workbook contains generated schema documentation for the selected tables.",
        "",
        "Included tables:",
    ]
    for line in lines:
        ws.append([line])
    for table_name in tables:
        ws.append([f"- {table_name}"])
    ws.freeze_panes = "A1"
    _auto_size_columns(ws)


def generate_excel_documentation(
    schema_map: Dict[str, List[Dict[str, Any]]],
    output_path: Path,
    connection_name: str,
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    create_readme_sheet(wb, connection_name, list(schema_map.keys()))
    create_table_index_sheet(wb, schema_map)
    create_table_relationships_sheet(wb, schema_map)
    for table_name, columns in schema_map.items():
        create_table_sheet(wb, table_name, columns)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
